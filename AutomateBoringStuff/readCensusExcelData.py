#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for
# each county.

import openpyxl, pprint, os, requests, zipfile
from pathlib import Path

# Check if file exists
if os.path.exists("automate_online-materials/censuspopdata.xlsx"):
    print("File exists.")
else:
    print("File doesn't exist, download starting...")

    # Gets the zip file containing all the downloads for the book
    file_url = "https://www.nostarch.com/download/Automate_the_Boring_Stuff_onlinematerials.zip"
    req = requests.get(file_url)

    with open("ATBSfiles.zip", 'wb') as f:
        for chunk in req.iter_content(chunk_size = 1024):
            f.write(chunk)

    print("Download Completed.")

    # Parse the zip file and extract the censuspopdata.xlsx file
    file = zipfile.ZipFile("ATBSfiles.zip")
    zipname = file.namelist()[0]
    file.extract(zipname + "censuspopdata.xlsx")

print('Opening workbook...')

wb = openpyxl.load_workbook("automate_online-materials/censuspopdata.xlsx")
sheet = wb['Population by Census Tract']
countyData = {}

# Fill in countyData with each county's population and tracts.
print('Reading data...', sheet.max_row)

for row in range(2, int(sheet.max_row) + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    # Make sure the key for this county in this state exists.
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    # Increase the county pop by the pop in this census tract.
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print('Writing results...')

resultFile = open('census2010Organized99.txt', 'w')
resultFile.write('all data = ' + pprint.pformat(countyData))
resultFile.close()

print('Done!!!')
